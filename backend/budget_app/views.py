from rest_framework import viewsets
from rest_framework import generics
from django.db import models


from .models import Prestataire
from .models import RIB
from .serializers import PrestataireSerializer
from .serializers import RIBSerializer
from .models import Budget
from .serializers import BudgetSerializer
from .models import LigneBudgetaire
from .serializers import LigneBudgetaireSerializer
from .models import Ordonnancement
from .serializers import OrdonnancementSerializer
from .models import Reglement
from .serializers import ReglementSerializer
from .models import CodeBanque, CodeLocalite
from .serializers import CodeBanqueSerializer, CodeLocaliteSerializer

class PrestataireViewSet(viewsets.ModelViewSet):
    queryset = Prestataire.objects.all()
    serializer_class = PrestataireSerializer


class RIBViewSet(viewsets.ModelViewSet):
    queryset = RIB.objects.all()
    serializer_class = RIBSerializer


class BudgetViewSet(viewsets.ModelViewSet):
    queryset = Budget.objects.all()
    serializer_class = BudgetSerializer


class DepensesBudgetViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Budget.objects.filter(type='depenses')
    serializer_class = BudgetSerializer

class LigneBudgetaireViewSet(viewsets.ModelViewSet):
    queryset = LigneBudgetaire.objects.all()
    serializer_class = LigneBudgetaireSerializer



class OrdonnancementViewSet(viewsets.ModelViewSet):
    queryset = Ordonnancement.objects.all()
    serializer_class = OrdonnancementSerializer


from rest_framework.response import Response
from rest_framework import status

from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from .models import Reglement, Ordonnancement, LigneBudgetaire
from .serializers import ReglementSerializer


from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import Reglement, Ordonnancement, LigneBudgetaire
from .serializers import ReglementSerializer

class ReglementViewSet(viewsets.ModelViewSet):
    queryset = Reglement.objects.all()
    serializer_class = ReglementSerializer

    # ➕ Création d’un règlement
    def create(self, request, *args, **kwargs):
        ligne_id = request.data.get('ligne_id')
        if not ligne_id:
            return Response({"detail": "ligne_id obligatoire"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Récupérer le premier ordonnancement lié à la ligne (éviter MultipleObjectsReturned)
        ordonnancement = Ordonnancement.objects.filter(ligne_id=ligne_id).first()
        if not ordonnancement:
            return Response({"detail": "Aucun ordonnancement trouvé pour cette ligne"}, status=status.HTTP_404_NOT_FOUND)
        
        data = request.data.copy()
        data['ordonnancement'] = ordonnancement.id
        
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        ligne = instance.ligne
        
        # Pour mettre à jour ou manipuler ordonnancements, utilise filter + first
        ordonnancement = Ordonnancement.objects.filter(ligne=ligne).first()
        if not ordonnancement:
            return Response({"detail": "Aucun ordonnancement trouvé pour cette ligne"}, status=status.HTTP_404_NOT_FOUND)
        
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    # ✏️ Modification d’un règlement
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()

        # On récupère les anciennes valeurs (au cas où tu veux en faire quelque chose, comme journalisation)
        old_montant = instance.montant_regler
        old_ligne = instance.ligne
        old_ordonnancement = instance.ordonnancement

        # Validation des nouvelles données
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        # Sauvegarde du nouveau règlement
        reglement = serializer.save()

        # Aucune modification manuelle sur les lignes ou ordonnancements n'est nécessaire ici.
        # Les totaux peuvent être recalculés dynamiquement si besoin via des agrégats.

        return Response(self.get_serializer(reglement).data, status=status.HTTP_200_OK)


    # ❌ Suppression d’un règlement
   


class CodeBanqueViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CodeBanque.objects.all()
    serializer_class = CodeBanqueSerializer

class CodeLocaliteViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CodeLocalite.objects.all()
    serializer_class = CodeLocaliteSerializer

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db import models
from decimal import Decimal
from .models import Budget, LigneBudgetaire, Ordonnancement, Reglement, Prestataire
def format_montant(montant):
    if montant is None:
        return ''
    
    # forcer en float
    montant = float(montant)
    
    
    s = f"{montant:,.2f}"  # ex: '1,234,567.89'
    
    # remplacer la virgule par espace (séparateur milliers)
    # et remplacer le point par virgule (décimal)
    s = s.replace(',', ' ').replace('.', ',')
    return s

def safe_str(value):
    """Retourne un espace si la valeur est vide ou invalide."""
    if value is None:
        return " "
    value_str = str(value).strip().lower()
    if value_str in ["", "nan"]:
        return " "
    return str(value)


def exporter_budget_excel(request, annee, type_budget):
    try:
        budget = Budget.objects.get(annee=annee, type=type_budget)
    except Budget.DoesNotExist:
        return HttpResponse("Budget pour l'année et type spécifiés non trouvé.", status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Budget {annee} - {type_budget}"

    colonnes = [
    "Rubrique",
    "Sous Rubrique",
    "Prestataire",
    "Type Prestataire",
    "Montant Budget TTC",
    "Montant Ordonnancé",
    "Montant Réglé",
    "Reliquat (Restant à régler)",
    "Commentaire",
    "Pourcentage Ordonnancement (%)",
    "Pourcentage Règlement (%)"
]
    ws.append(colonnes)

    lignes = LigneBudgetaire.objects.filter(budget=budget).select_related('prestataire')

    for ligne in lignes:
        montant_budget_ttc = ligne.montant_budget_ttc or Decimal('0')
        montant_ordonnancer = Ordonnancement.objects.filter(ligne=ligne).aggregate(
            total=models.Sum('montant_ordonnancement')
        )['total'] or Decimal('0')
        montant_regler = Reglement.objects.filter(ligne=ligne).aggregate(
            total=models.Sum('montant_regler')
        )['total'] or Decimal('0')
        reliquat = montant_budget_ttc - montant_regler
        pct_ordo_budget = (montant_ordonnancer / montant_budget_ttc) if montant_budget_ttc else Decimal('0')
        pct_regl_budget = (montant_regler / montant_budget_ttc) if montant_budget_ttc else Decimal('0')
        pct_ordo_budget *= 100
        pct_regl_budget *= 100
        pct_ordo_budget = f"{pct_ordo_budget:.2f} %"
        pct_regl_budget = f"{pct_regl_budget:.2f} %"
        prestataire = ligne.prestataire
        if prestataire.type == 'physique':
            nom_prestataire = f"{prestataire.nom or ''} {prestataire.prenom or ''}".strip()
        else:
            nom_prestataire = prestataire.raison_sociale or ''

        type_prestataire = prestataire.type
        commentaire = ""
        rubrique = safe_str(ligne.rubrique)
        sous_rubrique = safe_str(ligne.sous_rubrique)

        row = [
            rubrique,
            sous_rubrique,
            nom_prestataire,
            type_prestataire,
            format_montant(montant_budget_ttc),
            format_montant(montant_ordonnancer),
            format_montant(montant_regler),
            format_montant(reliquat),
            commentaire,
            pct_ordo_budget,
            pct_regl_budget,
        ]


        ws.append(row)

    for i, _ in enumerate(colonnes, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=budget_{annee}_{type_budget}.xlsx'

    wb.save(response)
    return response


import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework import status
from .models import Budget, LigneBudgetaire, Prestataire, Ordonnancement, Reglement
from datetime import date

class ImportBudgetView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        budget_id = request.POST.get('budget_id')
        annee = request.POST.get('annee')
        type_budget = request.POST.get('type_budget')

        if not file:
            return Response({"error": "Aucun fichier fourni."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file)
            df.columns = df.columns.str.strip()  # Nettoyer noms colonnes (supprimer espaces)
        except Exception as e:
            return Response({"error": f"Erreur lors de la lecture du fichier Excel: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        # Vérifier ou créer le budget
        if budget_id:
            try:
                budget = Budget.objects.get(id=budget_id)
            except Budget.DoesNotExist:
                return Response({"error": "Budget non trouvé."}, status=status.HTTP_404_NOT_FOUND)
        elif annee and type_budget:
            budget = Budget.objects.create(annee=annee, type=type_budget, date_creation=date.today())
        else:
            return Response({"error": "Veuillez fournir budget_id ou annee + type_budget."}, status=status.HTTP_400_BAD_REQUEST)

        excluded_labels = ['total', 'total général', 'total général:', 'totaux']

        for _, row in df.iterrows():
            rubrique = row.get('RUBRIQUE')
            sous_rubrique = row.get('SOUS RUBRIQUE')
            prestataire_nom = row.get('PRESTATAIRES')
            montant_budget = row.get('MONTANT BUDGET TTC 2025') or 0
            montant_ordonnancement = row.get('ORDONNANCEMENT') or 0
            montant_reglement = row.get('MONTANT REGLEMENT') or 0

            if pd.isna(prestataire_nom) or str(prestataire_nom).strip().lower() in excluded_labels or montant_budget == 0:
                continue

            prestataire, _ = Prestataire.objects.get_or_create(
                raison_sociale=prestataire_nom,
                defaults={
                    'type': 'morale',
                    'tel': '0000000000',
                    'adresse': 'Adresse inconnue',
                    'contact': 'Contact inconnu'
                }
            )

            ligne = LigneBudgetaire.objects.create(
                budget=budget,
                prestataire=prestataire,
                montant_budget_ttc=montant_budget,
                rubrique=rubrique,
                sous_rubrique=sous_rubrique,
            )

            if montant_ordonnancement > 0:
                Ordonnancement.objects.create(
                    ligne=ligne,
                    date_ordonnancement=date.today(),
                    montant_ordonnancement=montant_ordonnancement,
                    objet="Import Excel",
                    justificatifs="Import automatique",
                )

            if montant_reglement > 0:
                dernier_ord = ligne.ordonnancement_set.last()
                Reglement.objects.create(
                    ligne=ligne,
                    ordonnancement=dernier_ord,
                    date_paiement=date.today(),
                    montant_regler=montant_reglement,
                    justification="Import automatique",
                    compte_debite="00000000000000",
                    reglement_rel="Import automatique",
                    mode_paiement='virement'
                )

        return Response({"message": "Importation terminée avec succès."})

    

from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework import status
import pandas as pd
from .models import CodeBanque

class UpdateNomBanqueView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('domicialisation.xlsx')

        if not file:
            return Response({"error": "Aucun fichier fourni."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file)
            df.columns = df.columns.str.strip()  # Nettoie les noms de colonnes
        except Exception as e:
            return Response({"error": f"Erreur de lecture Excel : {e}"}, status=status.HTTP_400_BAD_REQUEST)

        updated = 0
        created = 0

        for _, row in df.iterrows():
            code = row.get('Code Banque')
            nom_banque = row.get('Nom_Banque')

            if pd.isna(code):
                continue

            code = str(int(code)) if isinstance(code, float) else str(code).strip()
            nom_banque = str(nom_banque).strip() if pd.notna(nom_banque) else ""

            obj, created_flag = CodeBanque.objects.update_or_create(
                code=code,
                defaults={'nom_banque': nom_banque}
            )

            if created_flag:
                created += 1
            else:
                updated += 1

        return Response({
            "message": "Mise à jour terminée.",
            "banques_mises_a_jour": updated,
            "banques_crees": created
        }, status=status.HTTP_200_OK)
